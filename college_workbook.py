import os
from openpyxl import load_workbook, Workbook
import requests
from string import ascii_uppercase as alc
from copy import copy
def create(type='enade',year='2021', courses=['CIÊNCIA DA COMPUTAÇÃO', 'TECNOLOGIA EM ANÁLISE E DESENVOLVIMENTO DE SISTEMAS', 'SISTEMAS DE INFORMAÇÃO', 'MEDICINA VETERINÁRIA']):
    #year = input("What year you want to check? \n")
    path_enade = f'./CollegeListBRconceito_enade{year}.xlsx'
    path_idd = f'./CollegeListBRconceito_idd{year}.xlsx'
    url_link = f'https://download.inep.gov.br/educacao_superior/indicadores/resultados/{year}'
    url_link_enade = f'{url_link}/conceito_enade_{year}.xlsx'
    url_link_idd = f'{url_link}/IDD_{year}.xlsx'
    if year == '2018':
        url_link_enade = 'https://download.inep.gov.br/educacao_superior/indicadores/legislacao/2019/resultados_conceito_enade_2018.xlsx'
        url_link_idd = 'https://download.inep.gov.br/educacao_superior/indicadores/legislacao/2019/resultados_IDD_2018.xlsx'
    elif year == '2019':
        url_link_enade = f'{url_link}/Conceito_Enade_2019.xlsx'
        url_link_idd = f'{url_link}/IDD_2019.xlsx'
    if not os.path.exists(path_enade):
        conceito_enade = requests.get(url_link_enade, verify=False)
        open(f"conceito_enade{year}.xlsx", "wb").write(conceito_enade.content)
    if not os.path.exists(path_idd):
        conceito_idd = requests.get(url_link_idd, verify=False)
        open(f"conceito_idd{year}.xlsx", "wb").write(conceito_idd.content)
    
    enade_wb = load_workbook(filename=f'conceito_enade{year}.xlsx')
    enade_ws2 = enade_wb.active
    dead_style = copy(enade_ws2['A1'].style)
    idd_wb = load_workbook(filename=f'conceito_idd{year}.xlsx')

    workbook = enade_wb
    filename = 'novo_conceito_enade.xlsx'
    if type == 'idd':
        filename = 'novo_conceito_idd.xlsx'
        workbook =  idd_wb
    enade_ws = workbook.active
    enade_array = []

    new_enade = Workbook()
    new_enade_ws = new_enade.active
    
    header_array = []
    new_enade_ws.row_dimensions[1] = enade_ws.row_dimensions[1]
    for i in alc:
        header_array.append(enade_ws[f'{i}1'].value)
        new_enade_ws.column_dimensions[i] = enade_ws.column_dimensions[i]

    new_enade_ws.append(header_array)

    for i in alc:
        new_enade_ws[f'{i}1'].font = copy(enade_ws[f'{i}1'].font)
        if year != '2018':
            new_enade_ws[f'{i}1'].style = copy(enade_ws[f'{i}1'].style)
        else:
            new_enade_ws[f'{i}1'].style = copy(dead_style)
        new_enade_ws[f'{i}1'].border = copy(enade_ws[f'{i}1'].border)
        new_enade_ws[f'{i}1'].alignment = copy(enade_ws[f'{i}1'].alignment)
    for data in enade_ws['C']:
        try:
            courses.index(data.value.strip())
            enade_array.append(data.row)
        except:
            pass
    for data in enade_array:
        cool_array = []
        for i in alc:
            cool_array.append(enade_ws[f'{i}{data}'].value)
        new_enade_ws.append(cool_array)
    new_enade.save(filename=filename)
    return filename
create()
create(type='idd')